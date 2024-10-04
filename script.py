import boto3
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# Função para obter todas as regiões disponíveis
def get_aws_regions():
    ec2 = boto3.client('ec2')
    regions = ec2.describe_regions()['Regions']
    return [region['RegionName'] for region in regions]

# Função principal para gerar o relatório por região
def generate_report(region):
    # Criando os clientes da AWS
    ec2_client = boto3.client('ec2', region_name=region)
    rds_client = boto3.client('rds', region_name=region)
    s3_client = boto3.client('s3', region_name=region)
    elb_client_v2 = boto3.client('elbv2', region_name=region)
    elb_client_v1 = boto3.client('elb', region_name=region)
    vpc_client = boto3.client('ec2', region_name=region)
    ecr_client = boto3.client('ecr', region_name=region)
    ecs_client = boto3.client('ecs', region_name=region)
    asg_client = boto3.client('autoscaling', region_name=region)
    eks_client = boto3.client('eks', region_name=region)
    ce_client = boto3.client('ce')
    sns_client = boto3.client('sns', region_name=region)

    # Obtendo o intervalo de datas para o custo
    end_date = datetime.utcnow().date()
    start_date = end_date - timedelta(days=30)  # Últimos 30 dias

    # Função para obter o custo de um serviço específico por recurso
    def get_cost_by_service(service_name):
        response = ce_client.get_cost_and_usage(
            TimePeriod={
                'Start': start_date.strftime('%Y-%m-%d'),
                'End': end_date.strftime('%Y-%m-%d')
            },
            Granularity='DAILY',
            Metrics=['UnblendedCost'],
            Filter={
                'Dimensions': {
                    'Key': 'SERVICE',
                    'Values': [service_name]
                }
            }
        )
        total_cost = sum(float(day['Total']['UnblendedCost']['Amount']) for day in response['ResultsByTime'])
        return total_cost

    # Coletando dados de EC2
    ec2_instances = ec2_client.describe_instances()
    ec2_data = []
    for reservation in ec2_instances['Reservations']:
        for instance in reservation['Instances']:
            instance_name = next((tag['Value'] for tag in instance.get('Tags', []) if tag['Key'] == 'Name'), 'N/A')
            ec2_data.append({
                'Serviço': 'EC2',
                'Nome do Recurso': instance_name,
                'ID': instance['InstanceId'],
                'Tipo': instance['InstanceType'],
                'Status': instance['State']['Name'],
                'Custo': get_cost_by_service('Amazon Elastic Compute Cloud')
            })

    # Coletando dados de RDS
    rds_instances = rds_client.describe_db_instances()
    rds_data = []
    for instance in rds_instances['DBInstances']:
        rds_data.append({
            'Serviço': 'RDS',
            'Nome do Recurso': instance['DBInstanceIdentifier'],
            'ID': instance['DBInstanceIdentifier'],
            'Tipo': instance['DBInstanceClass'],
            'Status': instance['DBInstanceStatus'],
            'Custo': get_cost_by_service('Amazon RDS')
        })

    # Coletando dados de S3
    s3_buckets = s3_client.list_buckets()
    s3_data = []
    for bucket in s3_buckets['Buckets']:
        # Obtém a região do bucket
        bucket_region = s3_client.get_bucket_location(Bucket=bucket['Name'])['LocationConstraint']
        if bucket_region is None:
            bucket_region = 'us-east-1'
        if bucket_region == region:
            s3_data.append({
                'Serviço': 'S3',
                'Nome do Recurso': bucket['Name'],
                'ID': bucket['Name'],
                'Tipo': 'Bucket',
                'Status': 'Ativo',
                'Custo': get_cost_by_service('Amazon Simple Storage Service')
            })

    # Coletando dados de ELB (v2)
    load_balancers_v2 = elb_client_v2.describe_load_balancers()
    elb_data_v2 = []
    for lb in load_balancers_v2['LoadBalancers']:
        elb_data_v2.append({
            'Serviço': 'ELB v2',
            'Nome do Recurso': lb['LoadBalancerName'],
            'ID': lb['LoadBalancerArn'],
            'Tipo': lb['Type'],
            'Status': lb['State']['Code'],
            'Custo': get_cost_by_service('Elastic Load Balancing v2')
        })

    # Coletando dados de ELB (v1)
    load_balancers_v1 = elb_client_v1.describe_load_balancers()
    elb_data_v1 = []
    for lb in load_balancers_v1['LoadBalancerDescriptions']:
        elb_data_v1.append({
            'Serviço': 'ELB v1',
            'Nome do Recurso': lb['LoadBalancerName'],
            'ID': lb['LoadBalancerName'],
            'Tipo': 'Classic Load Balancer',
            'Status': 'Ativo',
            'Custo': get_cost_by_service('Elastic Load Balancing')
        })

    # Coletando dados de NAT Gateways
    nat_gateways = vpc_client.describe_nat_gateways()
    nat_data = []
    for nat in nat_gateways['NatGateways']:
        nat_data.append({
            'Serviço': 'NAT Gateway',
            'Nome do Recurso': nat['NatGatewayId'],
            'ID': nat['NatGatewayId'],
            'Tipo': 'NAT Gateway',
            'Status': nat['State'],
            'Custo': get_cost_by_service('NAT Gateway')
        })

    # Coletando dados de ECR
    ecr_repositories = ecr_client.describe_repositories()
    ecr_data = []
    for repo in ecr_repositories['repositories']:
        ecr_data.append({
            'Serviço': 'ECR',
            'Nome do Recurso': repo['repositoryName'],
            'ID': repo['repositoryArn'],
            'Tipo': 'Repositório',
            'Status': 'Ativo',
            'Custo': get_cost_by_service('Amazon Elastic Container Registry')
        })

    # Coletando dados de ECS
    ecs_clusters = ecs_client.list_clusters()
    ecs_data = []
    for cluster_arn in ecs_clusters['clusterArns']:
        cluster_name = cluster_arn.split('/')[-1]
        services = ecs_client.list_services(cluster=cluster_arn)
        
        for service in services['serviceArns']:
            service_desc = ecs_client.describe_services(cluster=cluster_arn, services=[service])
            task_def_arn = service_desc['services'][0]['taskDefinition']
            task_def_desc = ecs_client.describe_task_definition(taskDefinition=task_def_arn)
            
            ecs_data.append({
                'Serviço': 'ECS',
                'Nome do Recurso': service_desc['services'][0]['serviceName'],
                'ID': service_desc['services'][0]['serviceArn'],
                'Tipo': 'Serviço',
                'Status': service_desc['services'][0]['status'],
                'Custo': get_cost_by_service('Amazon Elastic Container Service')
            })

            # Adicionando a definição da tarefa
            ecs_data.append({
                'Serviço': 'ECS',
                'Nome do Recurso': task_def_desc['taskDefinition']['family'],
                'ID': task_def_arn,
                'Tipo': 'Definição de Tarefa',
                'Status': 'Ativo',
                'Custo': get_cost_by_service('Amazon Elastic Container Service')
            })

    # Coletando dados de volumes EBS
    ebs_volumes = ec2_client.describe_volumes()
    ebs_data = []
    for volume in ebs_volumes['Volumes']:
        ebs_data.append({
            'Serviço': 'EBS',
            'Nome do Recurso': volume['VolumeId'],
            'ID': volume['VolumeId'],
            'Tipo': volume['VolumeType'],
            'Status': volume['State'],
            'Custo': get_cost_by_service('Amazon Elastic Block Store')
        })

    # Coletando dados de AMIs
    amis = ec2_client.describe_images(Owners=['self'])  # Filtrando pelas AMIs criadas pelo usuário
    ami_data = []
    for ami in amis['Images']:
        ami_data.append({
            'Serviço': 'AMI',
            'Nome do Recurso': ami['Name'],
            'ID': ami['ImageId'],
            'Tipo': 'AMI',
            'Status': 'Disponível',
            'Custo': 'N/A'  # AMIs não têm custo direto separado
        })

    # Coletando dados de Elastic IPs
    elastic_ips = ec2_client.describe_addresses()
    eip_data = []
    for eip in elastic_ips['Addresses']:
        eip_data.append({
            'Serviço': 'Elastic IP',
            'Nome do Recurso': eip.get('PublicIp', 'N/A'),
            'ID': eip.get('AllocationId', 'N/A'),
            'Tipo': 'Elastic IP',
            'Status': 'Ativo' if 'InstanceId' in eip else 'Não associado',
            'Custo': get_cost_by_service('Elastic IP')
        })

    # Coletando dados de Snapshots
    snapshots = ec2_client.describe_snapshots(OwnerIds=['self'])
    snapshot_data = []
    for snapshot in snapshots['Snapshots']:
        snapshot_data.append({
            'Serviço': 'Snapshot',
            'Nome do Recurso': snapshot['Description'],
            'ID': snapshot['SnapshotId'],
            'Tipo': 'Snapshot',
            'Status': snapshot['State'],
            'Custo': get_cost_by_service('Amazon Elastic Block Store')
        })

    # Coletando dados de Security Groups
    security_groups = ec2_client.describe_security_groups()
    sg_data = []
    for sg in security_groups['SecurityGroups']:
        sg_data.append({
            'Serviço': 'Security Group',
            'Nome do Recurso': sg['GroupName'],
            'ID': sg['GroupId'],
            'Tipo': 'Security Group',
            'Status': 'Ativo',
            'Custo': 'N/A'  # Security Groups não têm custo direto
        })

    # Coletando dados de VPCs
    vpcs = vpc_client.describe_vpcs()
    vpc_data = []
    for vpc in vpcs['Vpcs']:
        vpc_data.append({
            'Serviço': 'VPC',
            'Nome do Recurso': vpc['VpcId'],
            'ID': vpc['VpcId'],
            'Tipo': 'VPC',
            'Status': 'Ativo',
            'Custo': 'N/A'  # VPCs não têm custo direto
        })

    # Coletando dados de Subnets
    subnets = vpc_client.describe_subnets()
    subnet_data = []
    for subnet in subnets['Subnets']:
        subnet_data.append({
            'Serviço': 'Subnet',
            'Nome do Recurso': subnet['SubnetId'],
            'ID': subnet['SubnetId'],
            'Tipo': 'Subnet',
            'Status': 'Ativo',
            'Custo': 'N/A'  # Subnets não têm custo direto
        })

    # Coletando dados de Route Tables
    route_tables = vpc_client.describe_route_tables()
    route_table_data = []
    for rt in route_tables['RouteTables']:
        route_table_data.append({
            'Serviço': 'Route Table',
            'Nome do Recurso': rt['RouteTableId'],
            'ID': rt['RouteTableId'],
            'Tipo': 'Route Table',
            'Status': 'Ativo',
            'Custo': 'N/A'  # Route Tables não têm custo direto
        })

    # Coletando dados de Internet Gateways
    igws = vpc_client.describe_internet_gateways()
    igw_data = []
    for igw in igws['InternetGateways']:
        igw_data.append({
            'Serviço': 'Internet Gateway',
            'Nome do Recurso': igw['InternetGatewayId'],
            'ID': igw['InternetGatewayId'],
            'Tipo': 'Internet Gateway',
            'Status': 'Ativo',
            'Custo': 'N/A'  # IGWs não têm custo direto
        })

    # Coletando dados de Auto Scaling Groups
    asg_data = []
    asg_groups = asg_client.describe_auto_scaling_groups()
    for asg in asg_groups['AutoScalingGroups']:
        asg_data.append({
            'Serviço': 'Auto Scaling Group',
            'Nome do Recurso': asg['AutoScalingGroupName'],
            'ID': asg['AutoScalingGroupARN'],
            'Tipo': 'Auto Scaling Group',
            'Status': 'Ativo' if asg['HealthStatus'] == 'Healthy' else 'Não saudável',
            'Custo': get_cost_by_service('Amazon EC2 Auto Scaling')
        })

        # Coletando dados dos EC2 Instances em Auto Scaling Group
        for instance in asg['Instances']:
            asg_data.append({
                'Serviço': 'EC2 Instance (Auto Scaling)',
                'Nome do Recurso': instance.get('InstanceId', 'N/A'),
                'ID': instance['InstanceId'],
                'Tipo': 'EC2',
                'Status': instance['HealthStatus'],
                'Custo': get_cost_by_service('Amazon Elastic Compute Cloud')
            })

    # Coletando dados de EKS Clusters
    eks_data = []
    eks_clusters = eks_client.list_clusters()
    for cluster_name in eks_clusters['clusters']:
        cluster_info = eks_client.describe_cluster(name=cluster_name)['cluster']
        eks_data.append({
            'Serviço': 'EKS Cluster',
            'Nome do Recurso': cluster_info['name'],
            'ID': cluster_info['arn'],
            'Tipo': 'EKS',
            'Status': cluster_info['status'],
            'Custo': get_cost_by_service('Amazon Elastic Kubernetes Service')
        })

        # Coletando dados dos nós do EKS
        node_groups = eks_client.list_nodegroups(clusterName=cluster_name)
        for node_group in node_groups['nodegroups']:
            node_info = eks_client.describe_nodegroup(clusterName=cluster_name, nodegroupName=node_group)
            eks_data.append({
                'Serviço': 'EKS Node Group',
                'Nome do Recurso': node_info['nodegroup']['nodeGroupName'],
                'ID': node_info['nodegroup']['nodeGroupArn'],
                'Tipo': 'Node Group',
                'Status': node_info['nodegroup']['status'],
                'Custo': get_cost_by_service('Amazon Elastic Kubernetes Service')
            })

    # Unindo todos os dados
    all_data = (ec2_data + rds_data + s3_data + elb_data_v2 + elb_data_v1 +
                nat_data + ecr_data + ecs_data + ebs_data + ami_data +
                eip_data + snapshot_data + sg_data + vpc_data +
                subnet_data + route_table_data + igw_data + asg_data + eks_data)

    df = pd.DataFrame(all_data)

    # Salvando o relatório da região
    output_filename = f'relatorios/aws_cost_usage_report_{region}.xlsx'
    df.to_excel(output_filename, index=False)

    # Formatando a planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Custos"

    # Adicionando cabeçalhos
    headers = df.columns.tolist()
    ws.append(headers)

    # Definindo estilos para os cabeçalhos
    header_fill = PatternFill(start_color="00FFCC", end_color="00FFCC", fill_type="solid")
    bold_font = Font(bold=True)

    # Aplicando estilos aos cabeçalhos
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = bold_font

    # Adicionando os dados
    for row in df.itertuples(index=False):
        ws.append(row)

    # Salvando a planilha formatada
    wb.save(output_filename)

    print(f'Relatório gerado com sucesso: {output_filename}')

# Obtendo todas as regiões disponíveis
regions = get_aws_regions()

# Gerando o relatório para cada região
for region in regions:
    generate_report(region)

# Função global para gerar o relatório de IAM
def generate_global_iam_report():
    iam_client = boto3.client('iam')
    users = iam_client.list_users()
    iam_data = []

    for user in users['Users']:
        iam_data.append({
            'Serviço': 'IAM',
            'Nome do Recurso': user['UserName'],
            'ID': user['UserId'],
            'Tipo': 'Usuário',
            'Status': 'Ativo',
            'Custo': 'N/A'  # IAM não tem custo
        })

    df_iam = pd.DataFrame(iam_data)
    output_filename_iam = 'relatorios/aws_iam_report.xlsx'
    df_iam.to_excel(output_filename_iam, index=False)

    print(f'Relatório IAM gerado com sucesso: {output_filename_iam}')

# Gerando o relatório global de IAM
generate_global_iam_report()
